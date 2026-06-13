from __future__ import annotations

from dataclasses import asdict, dataclass, field
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_BAND_KEYS = ("BasicHours", "MonFriOvertime", "SatSunOvertime", "AnnualHoliday", "TotalPaidHours")
_EMP_AGENCY_ROWS = ("EMP", "AGENCY", "TOTAL")
_SHEET_LAST_COL = 8
_NUM_FORMAT = "0.00"
_PRIMARY_BLUE = "003078"

_MONTHLY_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FILL = PatternFill("solid", fgColor=_PRIMARY_BLUE)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_SECTION_TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="333333")
_SECTION_SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")
_SHEET_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=_PRIMARY_BLUE)

_BAND_LABELS = (
    "Basic hours",
    "Mon–Fri overtime",
    "Sat/Sun overtime",
    "Annual holiday",
    "Total paid hours",
)


@dataclass
class MonthlyEmployee:
    Name: str
    Category: str
    SageNo: int
    BasicHours: float
    MonFriOvertime: float
    SatSunOvertime: float
    AnnualHoliday: float
    TotalPaidHours: float
    IsHourly: bool = True


@dataclass
class MonthlyEmployeeTotal:
    Category: str
    BasicHours: float
    MonFriOvertime: float
    SatSunOvertime: float
    AnnualHoliday: float
    TotalPaidHours: float


@dataclass
class MonthlyAdjustment:
    Name: str
    Type: str
    Value: float


@dataclass
class MonthlyWeekSummary:
    employees: list[MonthlyEmployee] = field(default_factory=list)
    employee_totals: list[MonthlyEmployeeTotal] = field(default_factory=list)
    adjustments: list[MonthlyAdjustment] = field(default_factory=list)
    start_date: str = ""
    end_date: str = ""
    non_agency_total: float = 0.0
    grouped_totals: dict[str, MonthlyEmployeeTotal] = field(default_factory=dict)
    emp_agency_bands: dict[str, dict[str, float]] = field(default_factory=dict)


@dataclass
class EmployeeTableLayout:
    header_row: int
    data_start: int
    data_end: int
    next_row: int


@dataclass
class EmpAgencyLayout:
    header_row: int
    rows: dict[str, int]
    next_row: int


@dataclass
class WeekSheetLayout:
    sheet_name: str
    employee: EmployeeTableLayout
    emp_agency: EmpAgencyLayout


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _parse_decimal(value: Any) -> float:
    text = _to_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_int(value: Any) -> int:
    text = _to_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def _grouped_key(category: str) -> str:
    category = _to_text(category)
    if category.startswith("A-") and len(category) >= 9:
        return f"{category[:4]} {category[5:9]}"
    return category[:4] if len(category) >= 4 else category


def _is_agency_category(category: str) -> bool:
    return _to_text(category).upper().startswith("A-")


def _empty_bands() -> dict[str, float]:
    return {k: 0.0 for k in _BAND_KEYS}


def _compute_emp_agency_bands(employees: list[MonthlyEmployee]) -> dict[str, dict[str, float]]:
    emp = _empty_bands()
    agency = _empty_bands()
    for e in employees:
        target = agency if _is_agency_category(e.Category) else emp
        target["BasicHours"] += e.BasicHours
        target["MonFriOvertime"] += e.MonFriOvertime
        target["SatSunOvertime"] += e.SatSunOvertime
        target["AnnualHoliday"] += e.AnnualHoliday
        target["TotalPaidHours"] += e.TotalPaidHours
    total = {k: emp[k] + agency[k] for k in _BAND_KEYS}
    return {"EMP": emp, "AGENCY": agency, "TOTAL": total}


def _sum_emp_agency_bands(week_bands: list[dict[str, dict[str, float]]]) -> dict[str, dict[str, float]]:
    out = {row: _empty_bands() for row in _EMP_AGENCY_ROWS}
    for bands in week_bands:
        for row in _EMP_AGENCY_ROWS:
            for k in _BAND_KEYS:
                out[row][k] += float(bands.get(row, {}).get(k, 0.0) or 0.0)
    return out


def _employee_totals_from_employees(employees: list[MonthlyEmployee]) -> list[MonthlyEmployeeTotal]:
    by_cat: dict[str, MonthlyEmployeeTotal] = {}
    for e in employees:
        if e.Category not in by_cat:
            by_cat[e.Category] = MonthlyEmployeeTotal(e.Category, 0.0, 0.0, 0.0, 0.0, 0.0)
        t = by_cat[e.Category]
        t.BasicHours += e.BasicHours
        t.MonFriOvertime += e.MonFriOvertime
        t.SatSunOvertime += e.SatSunOvertime
        t.AnnualHoliday += e.AnnualHoliday
        t.TotalPaidHours += e.TotalPaidHours
    return list(by_cat.values())


def _build_grouped_totals(totals: list[MonthlyEmployeeTotal]) -> dict[str, MonthlyEmployeeTotal]:
    grouped: dict[str, MonthlyEmployeeTotal] = {}
    for t in totals:
        k = _grouped_key(t.Category)
        if k not in grouped:
            grouped[k] = MonthlyEmployeeTotal(k, 0.0, 0.0, 0.0, 0.0, 0.0)
        g = grouped[k]
        g.BasicHours += t.BasicHours
        g.MonFriOvertime += t.MonFriOvertime
        g.SatSunOvertime += t.SatSunOvertime
        g.AnnualHoliday += t.AnnualHoliday
        g.TotalPaidHours += t.TotalPaidHours
    return grouped


def _enrich_week_summary(out: MonthlyWeekSummary) -> None:
    if not out.employee_totals and out.employees:
        out.employee_totals = _employee_totals_from_employees(out.employees)
    out.non_agency_total = sum(t.TotalPaidHours for t in out.employee_totals if not _is_agency_category(t.Category))
    if not out.grouped_totals:
        out.grouped_totals = _build_grouped_totals(out.employee_totals)
    out.emp_agency_bands = _compute_emp_agency_bands(out.employees)


def _workbook_has_all_data(file_obj: Any) -> bool:
    file_obj.seek(0)
    try:
        xl = pd.ExcelFile(file_obj)
        return "All Data" in xl.sheet_names
    except Exception:
        return False
    finally:
        file_obj.seek(0)


def parse_weekly_gazebo_all_data(
    file_obj: Any,
    *,
    start_date: str = "",
    end_date: str = "",
) -> MonthlyWeekSummary:
    """Parse weekly Gazebo export (.xlsx) — employees from the All Data sheet."""
    file_obj.seek(0)
    df = pd.read_excel(file_obj, sheet_name="All Data", header=None, dtype=str)
    text_rows = [[_to_text(v) for v in row] for row in df.values.tolist()]
    out = MonthlyWeekSummary(start_date=start_date, end_date=end_date)
    if not text_rows:
        return out

    header_row = -1
    for i, row in enumerate(text_rows[:5]):
        if _to_text(row[0] if len(row) > 0 else "").lower() == "name" and _to_text(row[1] if len(row) > 1 else "").lower() == "category":
            header_row = i
            break
    if header_row < 0:
        return out

    for row in text_rows[header_row + 1 :]:
        name = _to_text(row[0] if len(row) > 0 else "")
        col_b = _to_text(row[1] if len(row) > 1 else "")
        if not name:
            if col_b.lower() in ("category", "category breakdown (overall)"):
                break
            break
        if col_b.startswith("Category breakdown"):
            break
        out.employees.append(
            MonthlyEmployee(
                Name=name,
                Category=col_b,
                SageNo=_parse_int(row[2] if len(row) > 2 else ""),
                BasicHours=_parse_decimal(row[3] if len(row) > 3 else ""),
                MonFriOvertime=_parse_decimal(row[4] if len(row) > 4 else ""),
                SatSunOvertime=_parse_decimal(row[5] if len(row) > 5 else ""),
                AnnualHoliday=_parse_decimal(row[6] if len(row) > 6 else ""),
                TotalPaidHours=_parse_decimal(row[7] if len(row) > 7 else ""),
            )
        )
    _enrich_week_summary(out)
    return out


def parse_monthly_week_file(file_obj: Any) -> MonthlyWeekSummary:
    from .payroll_service import _load_sheet  # reuse weekly reader

    table = _load_sheet(file_obj)
    rows = table.values.tolist()
    text_rows = [[_to_text(v) for v in row] for row in rows]
    if not text_rows:
        return MonthlyWeekSummary()

    out = MonthlyWeekSummary()
    out.start_date = _to_text(text_rows[0][4] if len(text_rows[0]) > 4 else "").removeprefix("D ").strip()
    out.end_date = _to_text(text_rows[0][7] if len(text_rows[0]) > 7 else "").removeprefix("D ").strip()

    r = 3  # row 4 in excel
    while r < len(text_rows):
        row = text_rows[r]
        if not _to_text(row[0] if len(row) > 0 else ""):
            break
        out.employees.append(
            MonthlyEmployee(
                Name=_to_text(row[0] if len(row) > 0 else ""),
                Category=_to_text(row[1] if len(row) > 1 else ""),
                SageNo=_parse_int(row[2] if len(row) > 2 else ""),
                BasicHours=_parse_decimal(row[3] if len(row) > 3 else ""),
                MonFriOvertime=_parse_decimal(row[4] if len(row) > 4 else ""),
                SatSunOvertime=_parse_decimal(row[5] if len(row) > 5 else ""),
                AnnualHoliday=_parse_decimal(row[6] if len(row) > 6 else ""),
                TotalPaidHours=_parse_decimal(row[7] if len(row) > 7 else ""),
            )
        )
        r += 1

    adjustments_header = -1
    for i in range(r, min(len(text_rows), r + 120)):
        if _to_text(text_rows[i][1] if len(text_rows[i]) > 1 else "").startswith("Adjustments"):
            adjustments_header = i
            break
    if adjustments_header >= 0:
        ar = adjustments_header + 2
        while ar < len(text_rows):
            name = _to_text(text_rows[ar][1] if len(text_rows[ar]) > 1 else "")
            if not name:
                break
            out.adjustments.append(
                MonthlyAdjustment(
                    Name=name,
                    Type=_to_text(text_rows[ar][2] if len(text_rows[ar]) > 2 else ""),
                    Value=_parse_decimal(text_rows[ar][3] if len(text_rows[ar]) > 3 else ""),
                )
            )
            ar += 1
        r = ar

    totals_header = -1
    for i in range(r, min(len(text_rows), r + 150)):
        if _to_text(text_rows[i][1] if len(text_rows[i]) > 1 else "") == "Category":
            totals_header = i
            break
    if totals_header >= 0:
        tr = totals_header + 1
        while tr < len(text_rows):
            cat = _to_text(text_rows[tr][1] if len(text_rows[tr]) > 1 else "")
            if not cat:
                break
            total = MonthlyEmployeeTotal(
                Category=cat,
                BasicHours=_parse_decimal(text_rows[tr][3] if len(text_rows[tr]) > 3 else ""),
                MonFriOvertime=_parse_decimal(text_rows[tr][4] if len(text_rows[tr]) > 4 else ""),
                SatSunOvertime=_parse_decimal(text_rows[tr][5] if len(text_rows[tr]) > 5 else ""),
                AnnualHoliday=_parse_decimal(text_rows[tr][6] if len(text_rows[tr]) > 6 else ""),
                TotalPaidHours=_parse_decimal(text_rows[tr][7] if len(text_rows[tr]) > 7 else ""),
            )
            out.employee_totals.append(total)
            tr += 1

    _enrich_week_summary(out)
    return out


def parse_monthly_inputs(
    weekly_files: list[Any],
    week_dates: list[tuple[str, str]] | None = None,
) -> list[MonthlyWeekSummary]:
    week_dates = week_dates or []
    summaries: list[MonthlyWeekSummary] = []
    for i, f in enumerate(weekly_files):
        start_date, end_date = week_dates[i] if i < len(week_dates) else ("", "")
        f.seek(0)
        if _workbook_has_all_data(f):
            f.seek(0)
            s = parse_weekly_gazebo_all_data(f, start_date=start_date, end_date=end_date)
        else:
            f.seek(0)
            s = parse_monthly_week_file(f)
            if start_date:
                s.start_date = start_date
            if end_date:
                s.end_date = end_date
        summaries.append(s)
    return summaries


def monthly_summaries_to_json(summaries: list[MonthlyWeekSummary]) -> list[dict[str, Any]]:
    return [asdict(s) for s in summaries]


def monthly_summaries_from_json(data: list[dict[str, Any]]) -> list[MonthlyWeekSummary]:
    out: list[MonthlyWeekSummary] = []
    for d in data:
        s = MonthlyWeekSummary(
            employees=[MonthlyEmployee(**e) for e in d.get("employees", [])],
            employee_totals=[MonthlyEmployeeTotal(**t) for t in d.get("employee_totals", [])],
            adjustments=[MonthlyAdjustment(**a) for a in d.get("adjustments", [])],
            start_date=str(d.get("start_date", "")),
            end_date=str(d.get("end_date", "")),
            non_agency_total=float(d.get("non_agency_total", 0.0)),
            emp_agency_bands=dict(d.get("emp_agency_bands") or {}),
        )
        s.grouped_totals = {}
        for k, v in (d.get("grouped_totals") or {}).items():
            if isinstance(v, dict):
                s.grouped_totals[str(k)] = MonthlyEmployeeTotal(**v)
        if not s.emp_agency_bands and s.employees:
            _enrich_week_summary(s)
        out.append(s)
    return out


def _set_column_widths(ws) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    for col in ("D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 12


def _apply_table_style(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    header_row: int | None = None,
) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = _MONTHLY_BORDER
            if header_row is not None and r == header_row:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif isinstance(cell.value, (int, float)) or (
                isinstance(cell.value, str) and str(cell.value).startswith("=")
            ):
                cell.number_format = _NUM_FORMAT
                cell.alignment = Alignment(horizontal="right", vertical="center")


def _write_sheet_banner(ws, title: str, start_date: str, end_date: str) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_SHEET_LAST_COL)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = _SHEET_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    period_parts: list[str] = []
    if start_date:
        period_parts.append(f"D {start_date}")
    if end_date:
        period_parts.append(f"D {end_date}")
    period_text = f"Period: {' to '.join(period_parts)}" if period_parts else ""
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_SHEET_LAST_COL)
    period_cell = ws.cell(2, 1, period_text)
    period_cell.font = _SECTION_SUBTITLE_FONT
    period_cell.alignment = Alignment(horizontal="left", vertical="center")

    _set_column_widths(ws)
    return 4


def _write_section_title(ws, r: int, title: str, subtitle: str = "") -> int:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_SHEET_LAST_COL)
    ws.cell(r, 1, title).font = _SECTION_TITLE_FONT
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")
    r += 1
    if subtitle:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_SHEET_LAST_COL)
        ws.cell(r, 1, subtitle).font = _SECTION_SUBTITLE_FONT
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1
    return r


def _band_col_index(band_index: int) -> int:
    """0-based index into _BAND_KEYS -> Excel column (D=4)."""
    return 4 + band_index


def _xl_sumifs_category(row: int, data_start: int, data_end: int, band_col: int) -> str:
    col = get_column_letter(band_col)
    if data_end < data_start:
        return "=0"
    return (
        f"=SUMIFS(${col}${data_start}:${col}${data_end},"
        f"$B${data_start}:$B${data_end},$B{row})"
    )


def _xl_sumproduct_agency(data_start: int, data_end: int, band_col: int) -> str:
    col = get_column_letter(band_col)
    if data_end < data_start:
        return "=0"
    return (
        f'=SUMPRODUCT((LEFT($B${data_start}:$B${data_end},2)="A-")*'
        f"(${col}${data_start}:${col}${data_end}))"
    )


def _xl_emp_band(data_start: int, data_end: int, band_col: int, agency_row: int) -> str:
    col = get_column_letter(band_col)
    if data_end < data_start:
        return "=0"
    return f"=SUM(${col}${data_start}:${col}${data_end})-{col}{agency_row}"


def _xl_total_band(emp_row: int, agency_row: int, band_col: int) -> str:
    col = get_column_letter(band_col)
    return f"={col}{emp_row}+{col}{agency_row}"


def _xl_cross_week_sumif(name_cell: str, week_sheet_names: list[str], band_col: int) -> str:
    col = get_column_letter(band_col)
    parts = [
        f"SUMIF('{sheet}'!$A:$A,{name_cell},'{sheet}'!${col}:${col})"
        for sheet in week_sheet_names
    ]
    return "=" + "+".join(parts)


def _xl_sum_week_emp_cells(week_layouts: list[WeekSheetLayout], row_key: str, band_col: int) -> str:
    col = get_column_letter(band_col)
    parts = [
        f"'{layout.sheet_name}'!{col}{layout.emp_agency.rows[row_key]}"
        for layout in week_layouts
    ]
    return "=" + "+".join(parts)


def _write_employee_header(ws, r: int) -> None:
    labels = (
        "Name",
        "Category",
        "Pay ID (Sage)",
        *_BAND_LABELS,
    )
    for i, label in enumerate(labels):
        ws.cell(r, 1 + i, label)


def _write_total_header(ws, r: int) -> None:
    ws.cell(r, 2, "Category")
    for j, label in enumerate(_BAND_LABELS):
        ws.cell(r, 4 + j, label)


def _write_adjustment_header(ws, r: int) -> None:
    ws.cell(r, 2, "Name")
    ws.cell(r, 3, "Type")
    ws.cell(r, 4, "Value")


def _write_employee_row(ws, r: int, e: MonthlyEmployee) -> None:
    ws.cell(r, 1, e.Name)
    ws.cell(r, 2, e.Category)
    ws.cell(r, 3, e.SageNo)
    ws.cell(r, 4, e.BasicHours)
    ws.cell(r, 5, e.MonFriOvertime)
    ws.cell(r, 6, e.SatSunOvertime)
    ws.cell(r, 7, e.AnnualHoliday)
    ws.cell(r, 8, e.TotalPaidHours)


def _write_total_row(ws, r: int, category: str, t: MonthlyEmployeeTotal) -> None:
    ws.cell(r, 2, category)
    ws.cell(r, 4, t.BasicHours)
    ws.cell(r, 5, t.MonFriOvertime)
    ws.cell(r, 6, t.SatSunOvertime)
    ws.cell(r, 7, t.AnnualHoliday)
    ws.cell(r, 8, t.TotalPaidHours)


def _write_emp_agency_block_formulas(
    ws,
    r: int,
    emp_layout: EmployeeTableLayout,
    *,
    row_label: str | None = None,
) -> tuple[int, dict[str, int]]:
    """Write EMP/AGENCY/TOTAL with SUMPRODUCT/SUM formulas from employee table."""
    block_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    emp_row = r
    agency_row = r + 1
    total_row = r + 2

    if row_label:
        ws.cell(r, 2, row_label)

    ws.cell(emp_row, 3, "EMP")
    ws.cell(agency_row, 3, "AGENCY")
    ws.cell(total_row, 3, "TOTAL")

    for j in range(5):
        col = _band_col_index(j)
        ws.cell(agency_row, col, _xl_sumproduct_agency(ds, de, col))
        ws.cell(emp_row, col, _xl_emp_band(ds, de, col, agency_row))
        ws.cell(total_row, col, _xl_total_band(emp_row, agency_row, col))

    row_map = {"EMP": emp_row, "AGENCY": agency_row, "TOTAL": total_row}
    _apply_table_style(ws, block_start, total_row, 2, _SHEET_LAST_COL)
    return total_row + 1, row_map


def _write_emp_agency_block(
    ws,
    r: int,
    bands: dict[str, dict[str, float]],
    *,
    row_label: str | None = None,
) -> int:
    """EMP/AGENCY/TOTAL in col C with bands in D–H; optional label in col B on first row."""
    block_start = r
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        if row_label and i == 0:
            ws.cell(r, 2, row_label)
        ws.cell(r, 3, key)
        row_bands = bands.get(key) or _empty_bands()
        for j, col in enumerate(_BAND_KEYS):
            ws.cell(r, 4 + j, float(row_bands.get(col, 0.0) or 0.0))
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    return r


def _write_emp_agency_section_formulas(
    ws,
    r: int,
    emp_layout: EmployeeTableLayout,
    section_title: str,
    subtitle: str,
    *,
    row_label: str | None = None,
) -> EmpAgencyLayout:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    r, row_map = _write_emp_agency_block_formulas(ws, r, emp_layout, row_label=row_label)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return EmpAgencyLayout(header_row=header_row, rows=row_map, next_row=r + 1)


def _write_emp_agency_section_monthly_formulas(
    ws,
    r: int,
    week_layouts: list[WeekSheetLayout],
    section_title: str,
    subtitle: str,
) -> tuple[int, EmpAgencyLayout]:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    block_start = r
    ws.cell(r, 2, "MONTHLY")
    row_map: dict[str, int] = {}
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        row_map[key] = r
        ws.cell(r, 3, key)
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_sum_week_emp_cells(week_layouts, key, _band_col_index(j)))
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1, EmpAgencyLayout(header_row=header_row, rows=row_map, next_row=r + 1)


def _write_emp_agency_block_week_refs(
    ws,
    r: int,
    week_layout: WeekSheetLayout,
    *,
    row_label: str | None = None,
) -> int:
    """Per-week block on Summary referencing Week sheet EMP/AGENCY/TOTAL cells."""
    block_start = r
    for i, key in enumerate(_EMP_AGENCY_ROWS):
        if row_label and i == 0:
            ws.cell(r, 2, row_label)
        ws.cell(r, 3, key)
        src_row = week_layout.emp_agency.rows[key]
        for j in range(5):
            col = _band_col_index(j)
            ws.cell(r, col, f"='{week_layout.sheet_name}'!{get_column_letter(col)}{src_row}")
        r += 1
    _apply_table_style(ws, block_start, r - 1, 2, _SHEET_LAST_COL)
    return r


def _write_emp_agency_section_diff_formulas(
    ws,
    r: int,
    summary_emp_layout: EmployeeTableLayout,
    monthly_layout: EmpAgencyLayout,
    section_title: str,
    subtitle: str,
) -> int:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    ds, de = summary_emp_layout.data_start, summary_emp_layout.data_end

    emp_merged_row = r
    ws.cell(emp_merged_row, 3, "EMP")
    agency_merged_row = r + 1
    ws.cell(agency_merged_row, 3, "AGENCY")
    total_merged_row = r + 2
    ws.cell(total_merged_row, 3, "TOTAL")
    ws.cell(emp_merged_row, 2, "Merged (employees)")

    for j in range(5):
        col = _band_col_index(j)
        ws.cell(agency_merged_row, col, _xl_sumproduct_agency(ds, de, col))
        ws.cell(emp_merged_row, col, _xl_emp_band(ds, de, col, agency_merged_row))
        ws.cell(total_merged_row, col, _xl_total_band(emp_merged_row, agency_merged_row, col))

    merged_rows = {"EMP": emp_merged_row, "AGENCY": agency_merged_row, "TOTAL": total_merged_row}
    diff_row_start = r + 3
    ws.cell(diff_row_start, 2, "Diff")

    for i, key in enumerate(_EMP_AGENCY_ROWS):
        row = diff_row_start + i
        ws.cell(row, 3, key)
        for j in range(5):
            col = _band_col_index(j)
            cl = get_column_letter(col)
            ws.cell(row, col, f"={cl}{merged_rows[key]}-{cl}{monthly_layout.rows[key]}")
        r = row + 1

    _apply_table_style(ws, header_row, r - 1, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_employee_table_cross_week(
    ws,
    r: int,
    employees: list[MonthlyEmployee],
    section_title: str,
    subtitle: str,
    week_sheet_names: list[str],
) -> EmployeeTableLayout:
    r = _write_section_title(
        ws,
        r,
        section_title,
        subtitle + " Totals use Excel formulas — click any hour cell to see Week sheet references.",
    )
    header_row = r
    _write_employee_header(ws, r)
    r += 1
    data_start = r
    for e in employees:
        ws.cell(r, 1, e.Name)
        ws.cell(r, 2, e.Category)
        ws.cell(r, 3, e.SageNo)
        name_ref = f"$A{r}"
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_cross_week_sumif(name_ref, week_sheet_names, _band_col_index(j)))
        r += 1
    data_end = r - 1 if employees else data_start - 1
    if employees:
        _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 1, _SHEET_LAST_COL, header_row=header_row)
    return EmployeeTableLayout(header_row=header_row, data_start=data_start, data_end=data_end, next_row=r + 1)


def _write_employee_table_source(
    ws,
    r: int,
    employees: list[MonthlyEmployee],
    section_title: str,
    subtitle: str,
) -> EmployeeTableLayout:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_employee_header(ws, r)
    r += 1
    data_start = r
    for e in employees:
        _write_employee_row(ws, r, e)
        r += 1
    data_end = r - 1 if employees else data_start - 1
    if employees:
        _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 1, _SHEET_LAST_COL, header_row=header_row)
    return EmployeeTableLayout(header_row=header_row, data_start=data_start, data_end=data_end, next_row=r + 1)


def _write_totals_table_formulas(
    ws,
    r: int,
    categories: list[str],
    emp_layout: EmployeeTableLayout,
    section_title: str,
    subtitle: str,
) -> int:
    if not categories:
        return r
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    data_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    for category in categories:
        ws.cell(r, 2, category)
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_sumifs_category(r, ds, de, _band_col_index(j)))
        r += 1
    _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_totals_table(
    ws,
    r: int,
    rows: list[tuple[str, MonthlyEmployeeTotal]],
    section_title: str,
    subtitle: str,
) -> int:
    if not rows:
        return r
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    for category, t in rows:
        _write_total_row(ws, r, category, t)
        r += 1
    _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_adjustments_table(ws, r: int, adjustments: list[MonthlyAdjustment]) -> int:
    r = _write_section_title(
        ws,
        r,
        f"Adjustments ({len(adjustments)})",
        "Manual hour corrections recorded in the legacy weekly file.",
    )
    header_row = r
    _write_adjustment_header(ws, r)
    r += 1
    for a in adjustments:
        ws.cell(r, 2, a.Name)
        ws.cell(r, 3, a.Type)
        ws.cell(r, 4, a.Value)
        r += 1
    if adjustments:
        _apply_table_style(ws, header_row, r - 1, 2, 4, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 2, 4, header_row=header_row)
    return r + 1


def _write_emp_agency_section(
    ws,
    r: int,
    bands: dict[str, dict[str, float]],
    section_title: str,
    subtitle: str,
    *,
    row_label: str | None = None,
) -> int:
    r = _write_section_title(ws, r, section_title, subtitle)
    header_row = r
    _write_total_header(ws, r)
    r += 1
    r = _write_emp_agency_block(ws, r, bands, row_label=row_label)
    _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def _write_summary_category_formulas(
    ws,
    r: int,
    categories: list[str],
    emp_layout: EmployeeTableLayout,
    merged_employees: dict[str, MonthlyEmployee],
) -> int:
    r = _write_section_title(
        ws,
        r,
        "Category totals (month)",
        "Sum of hours by work category — calculated with SUMIFS from the employee table above.",
    )
    header_row = r
    _write_total_header(ws, r)
    r += 1
    cat_data_start = r
    ds, de = emp_layout.data_start, emp_layout.data_end
    for category in categories:
        ws.cell(r, 2, category)
        for j in range(5):
            ws.cell(r, _band_col_index(j), _xl_sumifs_category(r, ds, de, _band_col_index(j)))
        r += 1
        non_hourly = [
            e for e in merged_employees.values()
            if not e.IsHourly and e.Category.upper() == category.upper()
        ]
        if non_hourly:
            ws.cell(r, 2, f"{category} non-hourly hours")
            ws.cell(r, 4, -sum(e.BasicHours for e in non_hourly))
            ws.cell(r, 5, -sum(e.MonFriOvertime for e in non_hourly))
            ws.cell(r, 6, -sum(e.SatSunOvertime for e in non_hourly))
            ws.cell(r, 7, -sum(e.AnnualHoliday for e in non_hourly))
            ws.cell(r, 8, -sum(e.TotalPaidHours for e in non_hourly))
            r += 1
    if r > cat_data_start:
        _apply_table_style(ws, header_row, r - 1, 1, _SHEET_LAST_COL, header_row=header_row)
    else:
        _apply_table_style(ws, header_row, header_row, 1, _SHEET_LAST_COL, header_row=header_row)
    return r + 1


def build_monthly_excel_bytes(
    week_summaries: list[MonthlyWeekSummary],
    non_hourly_names: set[str] | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    grouped_from_weeks: list[dict[str, MonthlyEmployeeTotal]] = []
    week_layouts: list[WeekSheetLayout] = []

    for i, s in enumerate(week_summaries, start=1):
        if not s.emp_agency_bands:
            _enrich_week_summary(s)
        sheet_name = f"Week{i}"
        ws = wb.create_sheet(sheet_name)
        r = _write_sheet_banner(ws, f"Gazebo HR — Week {i}", s.start_date, s.end_date)
        emp_layout = _write_employee_table_source(
            ws,
            r,
            s.employees,
            f"Week {i} — Employee paid hours",
            "One row per employee for this pay week (from weekly export).",
        )
        r = emp_layout.next_row
        r = _write_adjustments_table(ws, r, s.adjustments)
        categories = [t.Category for t in s.employee_totals]
        r = _write_totals_table_formulas(
            ws,
            r,
            categories,
            emp_layout,
            "Category totals",
            "Sum of hours by work category for this week (SUMIFS from employee rows above).",
        )
        grouped_rows = [(k, s.grouped_totals[k]) for k in sorted(s.grouped_totals.keys())]
        r = _write_totals_table(
            ws,
            r,
            grouped_rows,
            "Grouped totals (by category prefix)",
            "Categories rolled up to 4-character groups.",
        )
        emp_agency = _write_emp_agency_section_formulas(
            ws,
            r,
            emp_layout,
            "Gazebo vs agency summary",
            "EMP = Gazebo staff; AGENCY = A- categories; TOTAL = both (formulas from employee table).",
        )
        week_layouts.append(WeekSheetLayout(sheet_name=sheet_name, employee=emp_layout, emp_agency=emp_agency))
        grouped_from_weeks.append(s.grouped_totals)

    merged_employees: dict[str, MonthlyEmployee] = {}
    merged_totals: dict[str, MonthlyEmployeeTotal] = {}
    for s in week_summaries:
        for e in s.employees:
            cur = merged_employees.get(e.Name)
            if cur is None:
                merged_employees[e.Name] = MonthlyEmployee(**e.__dict__)
                continue
            cur.BasicHours += e.BasicHours
            cur.MonFriOvertime += e.MonFriOvertime
            cur.SatSunOvertime += e.SatSunOvertime
            cur.AnnualHoliday += e.AnnualHoliday
            cur.TotalPaidHours += e.TotalPaidHours
            cur.SageNo = e.SageNo
            cur.Category = e.Category

        for t in s.employee_totals:
            cur_t = merged_totals.get(t.Category)
            if cur_t is None:
                merged_totals[t.Category] = MonthlyEmployeeTotal(**t.__dict__)
                continue
            cur_t.BasicHours += t.BasicHours
            cur_t.MonFriOvertime += t.MonFriOvertime
            cur_t.SatSunOvertime += t.SatSunOvertime
            cur_t.AnnualHoliday += t.AnnualHoliday
            cur_t.TotalPaidHours += t.TotalPaidHours

    non_hourly_names = {n.strip().upper() for n in (non_hourly_names or set()) if n and n.strip()}
    for e in merged_employees.values():
        if e.Name.strip().upper() in non_hourly_names:
            e.IsHourly = False
        if not e.IsHourly and e.Category.strip().startswith("A-"):
            raise ValueError(f"Agency employee cannot be non-hourly: {e.Name}")

    week_sheet_names = [wl.sheet_name for wl in week_layouts]
    summary_start = week_summaries[0].start_date if week_summaries else ""
    summary_end = week_summaries[-1].end_date if week_summaries else ""
    ws = wb.create_sheet("Summary")
    r = _write_sheet_banner(ws, "Gazebo HR — Monthly Summary", summary_start, summary_end)
    summary_emp_layout = _write_employee_table_cross_week(
        ws,
        r,
        list(merged_employees.values()),
        "Monthly summary — all employees (weeks combined)",
        "Total paid hours per employee across all uploaded weeks.",
        week_sheet_names,
    )
    r = summary_emp_layout.next_row
    r = _write_summary_category_formulas(
        ws,
        r,
        list(merged_totals.keys()),
        summary_emp_layout,
        merged_employees,
    )

    agg_grouped: dict[str, MonthlyEmployeeTotal] = {}
    for m in grouped_from_weeks:
        for k, t in m.items():
            if k not in agg_grouped:
                agg_grouped[k] = MonthlyEmployeeTotal(k, 0.0, 0.0, 0.0, 0.0, 0.0)
            a = agg_grouped[k]
            a.BasicHours += t.BasicHours
            a.MonFriOvertime += t.MonFriOvertime
            a.SatSunOvertime += t.SatSunOvertime
            a.AnnualHoliday += t.AnnualHoliday
            a.TotalPaidHours += t.TotalPaidHours
    grouped_rows = [(k, agg_grouped[k]) for k in sorted(agg_grouped.keys())]
    r = _write_totals_table(
        ws,
        r,
        grouped_rows,
        "Grouped totals (month)",
        "Categories rolled up to 4-character groups across all weeks.",
    )

    if week_layouts:
        r, monthly_layout = _write_emp_agency_section_monthly_formulas(
            ws,
            r,
            week_layouts,
            "Month total — Gazebo vs agency",
            "Sum of weekly EMP/AGENCY/TOTAL rows across all uploaded weeks.",
        )

        r = _write_section_title(
            ws,
            r,
            "Per-week Gazebo vs agency",
            "Each row references the matching EMP/AGENCY/TOTAL cells on the Week sheet.",
        )
        for wi, layout in enumerate(week_layouts):
            label = "Weekly" if wi == 0 else None
            header_row = r
            _write_total_header(ws, r)
            r += 1
            r = _write_emp_agency_block_week_refs(ws, r, layout, row_label=label)
            _apply_table_style(ws, header_row, header_row, 2, _SHEET_LAST_COL, header_row=header_row)
            r += 1

        r = _write_emp_agency_section_diff_formulas(
            ws,
            r,
            summary_emp_layout,
            monthly_layout,
            "Reconciliation difference",
            "Merged (employees) minus MONTHLY row — EMP, AGENCY and TOTAL should all be zero.",
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
