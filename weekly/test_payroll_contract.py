"""Contract hours parsing: Clockrite report layout vs tabular XLS."""
from __future__ import annotations

import unittest
import zipfile
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from .payroll_service import (
    PayrollResult,
    _OVERALL_CATEGORY_ORDER,
    _build_analysis_dataframe,
    _build_grouped_analysis_dataframe,
    _build_overall_analysis_dataframe,
    _overall_category_key,
    audit_contract_pay_id_coverage,
    build_emp_agency_total_df,
    build_excel_bytes,
    calculate_payroll,
    parse_contracted_hours,
    parse_employee_display_names,
    parse_employee_hours,
    total_paid_hours_from_rows,
)


class TotalPaidHoursFromRowsTest(unittest.TestCase):
    def test_sums_including_a_prefix(self) -> None:
        rows = [
            {"TotalPaidHours": 4.0, "Category": "A-EL PROD"},
            {"TotalPaidHours": 3.5, "Category": "D-STAFF"},
        ]
        self.assertEqual(total_paid_hours_from_rows(rows), 7.5)


class ClockRitePaidHoursSummaryAnnualHLTest(unittest.TestCase):
    """Paid Hours (Inc Absence) Summary: Pay ID in B, Sage in D; annual duplicated at H and L (openpyxl col 8 and 12)."""

    def test_annual_from_excel_h_and_l_columns(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 2, "Pay ID")
        ws.cell(1, 4, "Sage")
        ws.cell(1, 5, "Hrs @ 1")
        ws.cell(1, 6, "Hrs @ 2")
        ws.cell(1, 8, "Hrs @ 3")
        ws.cell(1, 10, "Hrs @ 4")
        ws.cell(1, 12, "Hrs @ 5")
        ws.cell(2, 1, "TESTCAT")
        ws.cell(2, 2, "")
        ws.cell(3, 1, "Jane Doe")
        ws.cell(3, 2, 601)
        ws.cell(3, 3, 40)
        ws.cell(3, 4, 1.0)
        ws.cell(3, 5, 2.0)
        ws.cell(3, 8, 8.0)
        ws.cell(3, 12, 8.0)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        rows = parse_employee_hours(buf)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r["Category"], "TESTCAT")
        self.assertEqual(r["AnnualHoliday"], 8.0)
        self.assertEqual(r["BasicHours"], 32.0)
        self.assertEqual(r["MonFriOvertime"], 1.0)
        self.assertEqual(r["SatSunOvertime"], 2.0)
        self.assertEqual(r["TotalPaidHours"], 43.0)


class SagePayRefContractAliasTest(unittest.TestCase):
    """When Sage Pay Ref != Payroll Number in ClockRite block, Pay ID joins on Sage."""

    def test_sage_pay_ref_aliases_contract_hours_for_pay_id_join(self) -> None:
        wb = Workbook()
        ws = wb.active
        r = 0
        for label, val in (
            ("Contract Hrs", 41.25),
            ("Sage Pay Ref", 1528),
            ("Payroll Number", 1344),
        ):
            r += 1
            ws.cell(r, 6, label)
            ws.cell(r, 7, val)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        by_payroll, _ = parse_contracted_hours(buf)
        self.assertAlmostEqual(by_payroll[1344], 41.25, places=2)
        self.assertAlmostEqual(by_payroll[1528], 41.25, places=2)

    def test_calculate_payroll_joins_on_sage_no_when_payroll_number_differs(self) -> None:
        wb = Workbook()
        ws = wb.active
        r = 0
        for label, val in (
            ("Contract Hrs", 40.0),
            ("Sage Pay Ref", 1528),
            ("Payroll Number", 1344),
        ):
            r += 1
            ws.cell(r, 6, label)
            ws.cell(r, 7, val)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        employees = [
            {
                "SageNo": 1528,
                "Name": "S PATEL EOLL",
                "Category": "TEST",
                "TotalPaidHours": 45.0,
            },
        ]
        result = calculate_payroll(employees, buf)
        self.assertEqual(result.rows[0]["ContractHourMatch"], "Yes")
        self.assertEqual(result.rows[0]["ContractMatchReason"], "Matched on Pay ID")
        self.assertEqual(result.rows[0]["ContractedHours"], 40.0)
        self.assertEqual(result.rows[0]["Overtime"], 5.0)


class EmployeeDisplayNameLookupTest(unittest.TestCase):
    def _sonal_patel_block_workbook(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, 1344)
        ws.cell(1, 2, "Sonal Patel EOLL")
        ws.cell(2, 2, "Clock Name")
        ws.cell(2, 3, "S PATEL EOLL")
        ws.cell(3, 6, "Contract Hrs")
        ws.cell(3, 7, 40.0)
        ws.cell(4, 8, "Sage Pay Ref")
        ws.cell(4, 9, 1528)
        ws.cell(5, 6, "Payroll Number")
        ws.cell(5, 7, 1344)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_parse_display_names_maps_sage_and_clock(self) -> None:
        buf = self._sonal_patel_block_workbook()
        by_sage, by_clock = parse_employee_display_names(buf)
        self.assertEqual(by_sage[1528], "Sonal Patel EOLL")
        self.assertEqual(by_sage[1344], "Sonal Patel EOLL")
        self.assertEqual(by_clock["S PATEL EOLL"], "Sonal Patel EOLL")

    def test_calculate_payroll_replaces_clock_name_with_full_name(self) -> None:
        buf = self._sonal_patel_block_workbook()
        employees = [
            {
                "SageNo": 1528,
                "Name": "S PATEL EOLL",
                "Category": "TEST",
                "TotalPaidHours": 45.0,
            },
        ]
        result = calculate_payroll(employees, buf)
        self.assertEqual(result.rows[0]["Name"], "Sonal Patel EOLL")


class PayIdNetBasicTest(unittest.TestCase):
    def test_pay_id_path_nets_annual_from_basic(self) -> None:
        wb = Workbook()
        ws = wb.active
        # Header row: Pay ID in column C (0-based col 2)
        ws.cell(1, 2, "Name")
        ws.cell(1, 3, "Pay ID")
        ws.cell(1, 4, "Basic")
        ws.cell(1, 5, "MF")
        ws.cell(1, 6, "SS")
        ws.cell(1, 7, "Ann")
        # Category row (no numeric pay id)
        ws.cell(2, 1, "TESTCAT")
        ws.cell(2, 2, "")
        ws.cell(2, 3, "")
        # Employee: gross basic 40, annual 8 -> net basic 32, total still 40
        ws.cell(3, 1, "")
        ws.cell(3, 2, "Jane Doe")
        ws.cell(3, 3, 501)
        ws.cell(3, 4, 40)
        ws.cell(3, 5, 0)
        ws.cell(3, 6, 0)
        ws.cell(3, 7, 8)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        rows = parse_employee_hours(buf)
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r["Category"], "TESTCAT")
        self.assertEqual(r["BasicHours"], 32.0)
        self.assertEqual(r["AnnualHoliday"], 8.0)
        self.assertEqual(r["TotalPaidHours"], 40.0)


class EmpAgencyTotalDfTest(unittest.TestCase):
    def test_emp_plus_agency_equals_total(self) -> None:
        band = {"BasicHours": 10.0, "MonFriOvertime": 1.0, "SatSunOvertime": 2.0, "AnnualHoliday": 3.0, "TotalPaidHours": 16.0}
        gazebo = [{**band, "Category": "X", "Name": "A", "SageNo": 1}]
        agency = [{**band, "Category": "A-EL CLNR", "Name": "B", "SageNo": 2}]
        pr = PayrollResult(
            rows=gazebo + agency,
            agency_rows=agency,
            gazebo_rows=gazebo,
            total_paid_hours=32.0,
        )
        df = build_emp_agency_total_df(pr)
        self.assertEqual(len(df), 3)
        emp = df[df["Category"] == "EMP"].iloc[0]
        ag = df[df["Category"] == "AGENCY"].iloc[0]
        tot = df[df["Category"] == "TOTAL"].iloc[0]
        for col in ("BasicHours", "TotalPaidHours"):
            self.assertAlmostEqual(float(emp[col]) + float(ag[col]), float(tot[col]), places=5)


class BuildExcelNewSheetsTest(unittest.TestCase):
    def test_workbook_contains_new_sheet_names(self) -> None:
        band = {"BasicHours": 1.0, "MonFriOvertime": 0.0, "SatSunOvertime": 0.0, "AnnualHoliday": 0.0, "TotalPaidHours": 61.0}
        row = {
            **band,
            "Name": "HI",
            "Category": "C",
            "SageNo": 9,
            "ContractedHours": 0.0,
            "Overtime": 61.0,
            "ContractHourMatch": "No",
        }
        pr = PayrollResult(rows=[row], agency_rows=[], gazebo_rows=[row], total_paid_hours=61.0)
        data = build_excel_bytes(pr)
        with zipfile.ZipFile(BytesIO(data)) as zf:
            wbxml = zf.read("xl/workbook.xml").decode("utf-8")
        for name in ("EMP Agency Total", "Category summary", "Hours over 60"):
            self.assertIn(name, wbxml, msg=f"missing sheet {name!r}")

    def test_category_breakdown_block_column_offset(self) -> None:
        band = {
            "BasicHours": 10.0,
            "MonFriOvertime": 1.0,
            "SatSunOvertime": 2.0,
            "AnnualHoliday": 3.0,
            "TotalPaidHours": 16.0,
        }
        rows = [
            {**band, "Name": "A ONE", "Category": "CLNR", "SageNo": 1, "ContractedHours": 0.0, "Overtime": 0.0, "ContractHourMatch": "No"},
            {**band, "Name": "B TWO", "Category": "OFCE", "SageNo": 2, "ContractedHours": 0.0, "Overtime": 0.0, "ContractHourMatch": "No"},
        ]
        pr = PayrollResult(rows=rows, agency_rows=[], gazebo_rows=rows, total_paid_hours=32.0)
        data = build_excel_bytes(pr)
        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb["All Data"]
        self.assertEqual(ws.cell(len(rows) + 3, 2).value, "Category breakdown (detailed)")
        first_tier_row = len(rows) + 4
        tier_headers = [
            r
            for r in range(first_tier_row, ws.max_row + 1)
            if ws.cell(r, 2).value == "Category" and ws.cell(r, 4).value == "BasicHours"
        ]
        self.assertEqual(len(tier_headers), 4)
        header_row = tier_headers[0]
        self.assertIsNone(ws.cell(header_row, 1).value)
        self.assertEqual(ws.cell(header_row, 2).value, "Category")
        self.assertEqual(ws.cell(header_row, 4).value, "BasicHours")
        self.assertIsNone(ws.cell(header_row, 3).value)
        grand_row = header_row + 4
        self.assertEqual(ws.cell(grand_row, 2).value, "Grand total")
        self.assertEqual(ws.cell(grand_row, 4).value, 20.0)
        self.assertEqual(ws.cell(header_row, 2).border.left.style, "thin")

        grouped_header_row = tier_headers[1]
        self.assertEqual(ws.cell(grouped_header_row + 1, 2).value, "CLNR")

        emp_agency_df = build_emp_agency_total_df(pr)
        summary_header_row = tier_headers[2]
        self.assertEqual(ws.cell(summary_header_row, 2).value, "Category")
        for i, row in enumerate(emp_agency_df.itertuples(index=False)):
            r = summary_header_row + 1 + i
            self.assertEqual(ws.cell(r, 2).value, row.Category)
            self.assertEqual(ws.cell(r, 4).value, float(row.BasicHours))

        grand_tot_row = None
        diff_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 2).value == "GRAND TOTAL":
                grand_tot_row = r
            if ws.cell(r, 2).value == "Difference":
                diff_row = r
        self.assertIsNotNone(grand_tot_row)
        self.assertIsNotNone(diff_row)
        self.assertEqual(diff_row, grand_tot_row + 1)
        self.assertGreater(grand_tot_row, summary_header_row)
        self.assertEqual(ws.cell(grand_tot_row, 8).value, 32.0)
        for col in (4, 5, 6, 7, 8):
            self.assertEqual(ws.cell(diff_row, col).value, 0.0)
        wb.close()

    def test_all_data_rollup_totals_match(self) -> None:
        band = {
            "BasicHours": 10.0,
            "MonFriOvertime": 1.0,
            "SatSunOvertime": 2.0,
            "AnnualHoliday": 3.0,
            "TotalPaidHours": 16.0,
        }
        rows = [
            {**band, "Name": "A", "Category": "PROD BELT", "SageNo": 1, "ContractedHours": 0.0, "Overtime": 0.0, "ContractHourMatch": "No"},
            {**band, "Name": "B", "Category": "PKNG HIGH_RISK", "SageNo": 2, "ContractedHours": 0.0, "Overtime": 0.0, "ContractHourMatch": "No"},
        ]
        analysis = _build_analysis_dataframe(pd.DataFrame(rows))
        grouped = _build_grouped_analysis_dataframe(analysis)
        overall = _build_overall_analysis_dataframe(analysis)
        self.assertEqual(len(grouped), 2)
        self.assertEqual(list(overall["Category"]), list(_OVERALL_CATEGORY_ORDER))
        self.assertEqual(float(overall["TotalPaidHours"].sum()), 32.0)
        self.assertEqual(float(overall.loc[overall["Category"] == "PROD", "TotalPaidHours"].iloc[0]), 16.0)
        self.assertEqual(float(overall.loc[overall["Category"] == "PACK", "TotalPaidHours"].iloc[0]), 16.0)

    def test_overall_category_key_pack_includes_dpch(self) -> None:
        self.assertEqual(_overall_category_key("DPCH"), "PACK")
        self.assertEqual(_overall_category_key("A-EL DPCH"), "PACK")
        self.assertEqual(_overall_category_key("PKNG HIGH_RISK"), "PACK")
        self.assertEqual(_overall_category_key("A-EL PKNG SLEEVING"), "PACK")
        self.assertEqual(_overall_category_key("A-PM PKNG HIGH_RISK"), "PACK")
        self.assertEqual(_overall_category_key("A-RS PKNG SLEEVING"), "PACK")
        self.assertEqual(_overall_category_key("TECH TECHNICAL"), "TECH")
        self.assertEqual(_overall_category_key("A-EL TECHNICAL"), "TECH")
        self.assertEqual(_overall_category_key("OFCE"), "OFFICE")

    def test_overall_dpch_hours_roll_into_pack_on_fixture(self) -> None:
        with _OVERTIME_EMPLOYEE.open("rb") as ef, _OVERTIME_CONTRACT.open("rb") as cf:
            result = calculate_payroll(parse_employee_hours(ef), cf)
        analysis = _build_analysis_dataframe(pd.DataFrame(result.rows))
        overall = _build_overall_analysis_dataframe(analysis)
        dpch_hours = float(
            analysis.loc[analysis["Category"].isin(["DPCH", "A-EL DPCH"]), "TotalPaidHours"].sum()
        )
        ofce_hours = float(analysis.loc[analysis["Category"] == "OFCE", "TotalPaidHours"].sum())
        pack_hours = float(overall.loc[overall["Category"] == "PACK", "TotalPaidHours"].iloc[0])
        office_hours = float(overall.loc[overall["Category"] == "OFFICE", "TotalPaidHours"].iloc[0])
        self.assertGreater(dpch_hours, 0.0)
        self.assertAlmostEqual(office_hours, ofce_hours, places=2)
        self.assertGreaterEqual(pack_hours, dpch_hours)


_MONTH_DIR = Path(__file__).resolve().parent.parent / "data" / "month"
_GAZEBO_WEEKLY_XLSX = _MONTH_DIR / "gazebo_weekly_report_20260522-1401.xlsx"
_NEW_GAZEBO_WEEKLY_XLSX = _MONTH_DIR / "new_input_file/gazebo_weekly_report_20260605-1205.xlsx"
_MONTHLY_REF = _MONTH_DIR / "EXCEL_MONHTLY_MAY2026.xls"

_DATA = Path(__file__).resolve().parent.parent / "data"
_TEST_DATA = _DATA / "TEST_DATA"
_OVERTIME_TEST_DIR = _DATA / "ovettime_error_test_data"
_OVERTIME_EMPLOYEE = _OVERTIME_TEST_DIR / "dgross_paysummary2 (3).xls"
_OVERTIME_CONTRACT = _OVERTIME_TEST_DIR / "demployees_2023 (1).xls"
_CLOCKRITE = _DATA / "Employee contract hours - clockrite.xls"
_EMPLOYEE = _DATA / "dgross_paysummary2.xls"


@unittest.skipUnless(
    _OVERTIME_EMPLOYEE.is_file() and _OVERTIME_CONTRACT.is_file(),
    "data/ovettime_error_test_data fixtures not in repo",
)
class AllDataMultiTierIntegrationTest(unittest.TestCase):
    def test_grouped_and_overall_rollups_on_fixture(self) -> None:
        with _OVERTIME_EMPLOYEE.open("rb") as ef, _OVERTIME_CONTRACT.open("rb") as cf:
            result = calculate_payroll(parse_employee_hours(ef), cf)
        analysis = _build_analysis_dataframe(pd.DataFrame(result.rows))
        grouped = _build_grouped_analysis_dataframe(analysis)
        overall = _build_overall_analysis_dataframe(analysis)
        self.assertEqual(len(analysis), 26)
        self.assertEqual(len(grouped), 16)
        granular_total = float(analysis["TotalPaidHours"].sum())
        self.assertEqual(float(grouped["TotalPaidHours"].sum()), granular_total)
        self.assertEqual(float(overall["TotalPaidHours"].sum()), granular_total)
        self.assertEqual(list(overall["Category"]), list(_OVERALL_CATEGORY_ORDER))

        data = build_excel_bytes(result)
        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb["All Data"]
        first_tier_row = len(result.rows) + 4
        tier_headers = [
            r
            for r in range(first_tier_row, ws.max_row + 1)
            if ws.cell(r, 2).value == "Category" and ws.cell(r, 4).value == "BasicHours"
        ]
        self.assertEqual(len(tier_headers), 4)
        grand_tot_row = next(
            r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "GRAND TOTAL"
        )
        diff_row = next(
            r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "Difference"
        )
        self.assertEqual(diff_row, grand_tot_row + 1)
        self.assertEqual(ws.cell(grand_tot_row, 8).value, granular_total)
        for col in (4, 5, 6, 7, 8):
            self.assertAlmostEqual(float(ws.cell(diff_row, col).value or 0), 0.0, places=5)
        wb.close()


@unittest.skipUnless(
    _OVERTIME_EMPLOYEE.is_file() and _OVERTIME_CONTRACT.is_file(),
    "data/ovettime_error_test_data fixtures not in repo",
)
class ClockRiteOvertimeColumnRegressionTest(unittest.TestCase):
    def test_colleague_paysummary_747_and_579(self) -> None:
        with _OVERTIME_EMPLOYEE.open("rb") as f:
            rows = {r["SageNo"]: r for r in parse_employee_hours(f)}
        self.assertEqual(rows[747]["MonFriOvertime"], 5.5)
        self.assertEqual(rows[747]["TotalPaidHours"], 46.75)
        self.assertEqual(rows[579]["MonFriOvertime"], 9.25)
        self.assertEqual(rows[579]["SatSunOvertime"], 8.5)
        self.assertEqual(rows[579]["TotalPaidHours"], 17.75)

    def test_colleague_full_name_for_sage_1528(self) -> None:
        with _OVERTIME_EMPLOYEE.open("rb") as ef, _OVERTIME_CONTRACT.open("rb") as cf:
            employees = parse_employee_hours(ef)
            result = calculate_payroll(employees, cf)
        by_sage = {r["SageNo"]: r["Name"] for r in result.rows}
        self.assertEqual(by_sage.get(1528), "Sonal Patel EOLL")


@unittest.skipUnless(
    _OVERTIME_EMPLOYEE.is_file() and _OVERTIME_CONTRACT.is_file(),
    "data/ovettime_error_test_data fixtures not in repo",
)
class ContractMatchAuditTest(unittest.TestCase):
    """Audit: Contract match No = Pay ID missing from contract export (not a parser bug)."""

    _MISSING_PAY_IDS = {752, 1653, 1658, 1664, 1665}

    def test_five_pay_ids_missing_from_contract_export(self) -> None:
        with _OVERTIME_EMPLOYEE.open("rb") as ef, _OVERTIME_CONTRACT.open("rb") as cf:
            employees = parse_employee_hours(ef)
            missing = audit_contract_pay_id_coverage(employees, cf)
        missing_ids = {m["SageNo"] for m in missing}
        self.assertEqual(missing_ids, self._MISSING_PAY_IDS)

    def test_no_match_rows_have_reason_and_201_of_206_yes(self) -> None:
        with _OVERTIME_EMPLOYEE.open("rb") as ef, _OVERTIME_CONTRACT.open("rb") as cf:
            result = calculate_payroll(parse_employee_hours(ef), cf)
        no_rows = [r for r in result.rows if r["ContractHourMatch"] == "No"]
        self.assertEqual(len(no_rows), 5)
        self.assertEqual(sum(1 for r in result.rows if r["ContractHourMatch"] == "Yes"), 201)
        for r in no_rows:
            self.assertEqual(r["ContractMatchReason"], "Pay ID not in contract export")
            self.assertIn(r["SageNo"], self._MISSING_PAY_IDS)

    def test_after_contract_reexport_expect_yes_for_missing_ids(self) -> None:
        """Manual: re-export demployees from ClockRite with Pay IDs 752, 1653, 1658, 1664, 1665, then re-run this test."""
        self.assertTrue(
            _OVERTIME_CONTRACT.is_file(),
            "Re-export Employee Details (Advanced) .xls after adding missing staff in ClockRite",
        )


@unittest.skipUnless(
    _TEST_DATA.is_dir() and any(_TEST_DATA.glob("*.xls")),
    "data/TEST_DATA/*.xls not in repo",
)
class ClockRiteTestDataFilesTest(unittest.TestCase):
    def test_paid_hours_summary_files_parse_without_error(self) -> None:
        for path in sorted(_TEST_DATA.glob("*.xls")):
            with path.open("rb") as f:
                rows = parse_employee_hours(f)
            self.assertGreater(len(rows), 10, msg=f"{path.name} expected many employee rows")

            r0 = rows[0]
            for key in ("BasicHours", "AnnualHoliday", "TotalPaidHours", "MonFriOvertime", "SatSunOvertime"):
                self.assertIn(key, r0)


@unittest.skipUnless(_CLOCKRITE.is_file() and _EMPLOYEE.is_file(), "data/*.xls fixtures not in repo")
class ClockriteContractParseTest(unittest.TestCase):
    def test_parse_clockrite_export_maps_payroll_number_to_contract_hours(self) -> None:
        with _CLOCKRITE.open("rb") as f:
            by_payroll, _ = parse_contracted_hours(f)
        self.assertGreater(
            len(by_payroll),
            100,
            "Expected Clockrite 'Employee Details' report to yield many payroll keys",
        )
        # Known block from sample file: Payroll Number 1026, Contract Hrs 0
        self.assertIn(1026, by_payroll)
        self.assertEqual(by_payroll[1026], 0.0)

    def test_join_employee_pay_id_to_contract_payroll_number(self) -> None:
        with _EMPLOYEE.open("rb") as ef, _CLOCKRITE.open("rb") as cf:
            employees = parse_employee_hours(ef)
            result = calculate_payroll(employees, cf)
        matched = sum(1 for r in result.rows if r["ContractHourMatch"] == "Yes")
        # Sample data: 189/190 Pay IDs exist as Payroll Number in the contract file
        self.assertGreaterEqual(matched, 180)
        self.assertLess(matched, len(result.rows) + 1)


@unittest.skipUnless(_GAZEBO_WEEKLY_XLSX.is_file(), "data/month/gazebo weekly xlsx not in repo")
class MonthlyGazeboAllDataTest(unittest.TestCase):
    def test_parse_all_data_sheet(self) -> None:
        from .monthly_service import parse_weekly_gazebo_all_data

        with _GAZEBO_WEEKLY_XLSX.open("rb") as f:
            s = parse_weekly_gazebo_all_data(f, start_date="20/04/2026", end_date="26/04/2026")
        self.assertEqual(s.start_date, "20/04/2026")
        self.assertGreater(len(s.employees), 200)
        self.assertEqual(len(s.employee_totals), 26)
        self.assertIn("EMP", s.emp_agency_bands)
        self.assertAlmostEqual(s.emp_agency_bands["TOTAL"]["TotalPaidHours"], 5034.75, places=2)

    def test_build_monthly_workbook_summary_blocks(self) -> None:
        from .monthly_service import build_monthly_excel_bytes, parse_weekly_gazebo_all_data

        with _GAZEBO_WEEKLY_XLSX.open("rb") as f:
            week = parse_weekly_gazebo_all_data(f)
        summaries = [week, week, week, week]
        data = build_monthly_excel_bytes(summaries)
        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb["Summary"]
        labels = {ws.cell(r, 2).value for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value}
        self.assertIn("MONTHLY", labels)
        self.assertIn("Weekly", labels)
        self.assertIn("Diff", labels)
        monthly_emp_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "MONTHLY")
        self.assertEqual(ws.cell(monthly_emp_row, 3).value, "EMP")
        monthly_total_formula = str(ws.cell(monthly_emp_row, 8).value)
        self.assertTrue(monthly_total_formula.startswith("="))
        self.assertIn("'Week1'!", monthly_total_formula)
        self.assertIn("'Week4'!", monthly_total_formula)
        cat_formula = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 4).value
            if isinstance(val, str) and val.startswith("=SUMIFS"):
                cat_formula = val
                break
        self.assertIsNotNone(cat_formula, "Expected SUMIFS formula in category totals")
        emp_formula = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 4).value
            if isinstance(val, str) and "SUMIF('Week1'!" in val:
                emp_formula = val
                break
        self.assertIsNotNone(emp_formula, "Expected cross-week SUMIF on Summary employee table")
        diff_emp_row = next(
            r for r in range(1, ws.max_row + 1)
            if ws.cell(r, 2).value == "Diff" and ws.cell(r, 3).value == "EMP"
        )
        diff_total_row = diff_emp_row + 2
        self.assertEqual(ws.cell(diff_total_row, 3).value, "TOTAL")
        diff_total_formula = str(ws.cell(diff_total_row, 8).value)
        self.assertIn("-H", diff_total_formula)
        self.assertNotIn("'Week1'!", diff_total_formula, "Diff TOTAL should reference MONTHLY row, not raw week sum")
        monthly_emp_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "MONTHLY")
        monthly_total_row = monthly_emp_row + 2
        self.assertEqual(ws.cell(monthly_total_row, 3).value, "TOTAL")
        self.assertIn(f"H{monthly_total_row}", diff_total_formula)
        wb.close()

    def test_week_sheet_has_emp_agency_in_col_c(self) -> None:
        from .monthly_service import build_monthly_excel_bytes, parse_weekly_gazebo_all_data

        with _GAZEBO_WEEKLY_XLSX.open("rb") as f:
            week = parse_weekly_gazebo_all_data(f)
        data = build_monthly_excel_bytes([week])
        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb["Week1"]
        emp_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 3).value == "EMP")
        emp_basic = str(ws.cell(emp_row, 4).value)
        self.assertTrue(emp_basic.startswith("=SUM("))
        agency_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 3).value == "AGENCY")
        agency_basic = str(ws.cell(agency_row, 4).value)
        self.assertTrue(agency_basic.startswith("=SUMPRODUCT"))
        cat_row = next(
            r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 4).value, str) and str(ws.cell(r, 4).value).startswith("=SUMIFS")
        )
        self.assertIn("$B$", str(ws.cell(cat_row, 4).value))
        wb.close()


@unittest.skipUnless(_NEW_GAZEBO_WEEKLY_XLSX.is_file(), "data/month/new_input_file fixture not in repo")
class NewGazeboWeeklyInputTest(unittest.TestCase):
    def test_parse_new_weekly_export(self) -> None:
        from .monthly_service import parse_weekly_gazebo_all_data

        with _NEW_GAZEBO_WEEKLY_XLSX.open("rb") as f:
            s = parse_weekly_gazebo_all_data(f, start_date="01/06/2026", end_date="07/06/2026")
        self.assertEqual(len(s.employees), 190)
        self.assertEqual(s.employees[-1].Name, "SURENDRAN SUBRAMANI")
        self.assertAlmostEqual(s.emp_agency_bands["TOTAL"]["TotalPaidHours"], 5557.75, places=2)

    def test_build_monthly_workbook_from_new_export(self) -> None:
        from .monthly_service import build_monthly_excel_bytes, parse_weekly_gazebo_all_data

        with _NEW_GAZEBO_WEEKLY_XLSX.open("rb") as f:
            week = parse_weekly_gazebo_all_data(f)
        data = build_monthly_excel_bytes([week, week, week, week])
        wb = load_workbook(BytesIO(data), read_only=True)
        ws = wb["Summary"]
        labels = {ws.cell(r, 2).value for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value}
        self.assertIn("MONTHLY", labels)
        self.assertIn("Diff", labels)
        monthly_emp_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value == "MONTHLY")
        monthly_total_formula = str(ws.cell(monthly_emp_row, 8).value)
        self.assertTrue(monthly_total_formula.startswith("="))
        self.assertIn("'Week1'!", monthly_total_formula)
        self.assertIn("'Week4'!", monthly_total_formula)
        wb.close()


@unittest.skipUnless(_MONTHLY_REF.is_file(), "data/month/EXCEL_MONHTLY_MAY2026.xls not in repo")
class MonthlyLegacyWeekParseTest(unittest.TestCase):
    def test_legacy_w1_still_parses(self) -> None:
        from .monthly_service import parse_monthly_week_file

        with (_MONTHLY_REF).open("rb") as f:
            s = parse_monthly_week_file(f)
        self.assertGreater(len(s.employees), 50)
        self.assertIn("EMP", s.emp_agency_bands)
