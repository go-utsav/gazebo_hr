"""Parity tests for the Monthly service vs the legacy .NET Monthly.exe behaviour."""
from __future__ import annotations

import unittest
import zipfile
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook

from .monthly_service import (
    CategoryConflict,
    MonthlyAdjustment,
    MonthlyEmployee,
    MonthlyEmployeeTotal,
    MonthlyWeekSummary,
    _grouped_key,
    _parse_decimal,
    _parse_int,
    build_monthly_excel_bytes,
    detect_category_conflicts,
    monthly_summaries_from_json,
    monthly_summaries_to_json,
    parse_monthly_week_file,
)


_ZERO = Decimal("0")


def _emp(name: str, sage: int, cat: str, basic: str = "0", mf: str = "0", ss: str = "0", ah: str = "0", tp: str | None = None, hourly: bool = True) -> MonthlyEmployee:
    b = Decimal(basic)
    m = Decimal(mf)
    s = Decimal(ss)
    a = Decimal(ah)
    t = Decimal(tp) if tp is not None else (b + m + s + a)
    return MonthlyEmployee(
        Name=name, Category=cat, SageNo=sage,
        BasicHours=b, MonFriOvertime=m, SatSunOvertime=s,
        AnnualHoliday=a, TotalPaidHours=t, IsHourly=hourly,
    )


def _tot(cat: str, basic: str = "0", mf: str = "0", ss: str = "0", ah: str = "0", tp: str | None = None) -> MonthlyEmployeeTotal:
    b = Decimal(basic)
    m = Decimal(mf)
    s = Decimal(ss)
    a = Decimal(ah)
    t = Decimal(tp) if tp is not None else (b + m + s + a)
    return MonthlyEmployeeTotal(Category=cat, BasicHours=b, MonFriOvertime=m, SatSunOvertime=s, AnnualHoliday=a, TotalPaidHours=t)


class DecimalPrecisionTest(unittest.TestCase):
    """1000 × 0.01 must equal exactly 10.00 — float would drift, decimal must not."""

    def test_thousand_cent_increments_are_exact(self) -> None:
        e = _emp("ALICE", 1, "PROD", basic="0.00")
        for _ in range(1000):
            e.BasicHours += Decimal("0.01")
        self.assertEqual(e.BasicHours, Decimal("10.00"))
        self.assertEqual(str(e.BasicHours), "10.00")

    def test_one_third_repeating_does_not_drift(self) -> None:
        s = _ZERO
        for _ in range(300):
            s += Decimal("0.1") + Decimal("0.2")
        self.assertEqual(s, Decimal("90.0"))


class ParseHelperTest(unittest.TestCase):
    def test_parse_decimal_returns_decimal(self) -> None:
        self.assertEqual(_parse_decimal("1.50"), Decimal("1.50"))
        self.assertEqual(_parse_decimal("1,234.56"), Decimal("1234.56"))
        self.assertEqual(_parse_decimal(""), _ZERO)
        self.assertEqual(_parse_decimal(None), _ZERO)
        self.assertEqual(_parse_decimal("garbage"), _ZERO)

    def test_parse_int_handles_decimal_text(self) -> None:
        self.assertEqual(_parse_int("123"), 123)
        self.assertEqual(_parse_int("123.0"), 123)
        self.assertEqual(_parse_int(""), 0)
        self.assertEqual(_parse_int("garbage"), 0)


class GroupedKeyTest(unittest.TestCase):
    def test_agency_key_uses_split_5_to_9(self) -> None:
        self.assertEqual(_grouped_key("A-EL CLNR"), "A-EL CLNR")
        self.assertEqual(_grouped_key("A-EL PROD BELT"), "A-EL PROD")

    def test_short_agency_falls_back_to_first_four(self) -> None:
        self.assertEqual(_grouped_key("A-X"), "A-X")

    def test_non_agency_key_is_first_four(self) -> None:
        self.assertEqual(_grouped_key("D-PROD FRYING"), "D-PR")
        self.assertEqual(_grouped_key("ABC"), "ABC")


class CategoryConflictTest(unittest.TestCase):
    def test_same_sageno_different_categories_flags_conflict(self) -> None:
        s1 = MonthlyWeekSummary(employees=[_emp("ALICE", 501, "D-PROD")])
        s2 = MonthlyWeekSummary(employees=[_emp("ALICE", 501, "D-PACK")])
        conflicts = detect_category_conflicts([s1, s2])
        self.assertEqual(len(conflicts), 1)
        c = conflicts[0]
        self.assertEqual(c.sage_no, 501)
        self.assertEqual(c.name, "ALICE")
        self.assertEqual(sorted(c.categories), ["D-PACK", "D-PROD"])

    def test_consistent_category_no_conflict(self) -> None:
        s1 = MonthlyWeekSummary(employees=[_emp("ALICE", 501, "D-PROD")])
        s2 = MonthlyWeekSummary(employees=[_emp("ALICE", 501, "D-PROD")])
        self.assertEqual(detect_category_conflicts([s1, s2]), [])

    def test_zero_sageno_skipped(self) -> None:
        s1 = MonthlyWeekSummary(employees=[_emp("BOB", 0, "X")])
        s2 = MonthlyWeekSummary(employees=[_emp("CAROL", 0, "Y")])
        self.assertEqual(detect_category_conflicts([s1, s2]), [])

    def test_categories_compared_case_insensitive(self) -> None:
        s1 = MonthlyWeekSummary(employees=[_emp("ALICE", 501, "d-prod")])
        s2 = MonthlyWeekSummary(employees=[_emp("ALICE", 501, "D-PROD")])
        self.assertEqual(detect_category_conflicts([s1, s2]), [])


class JsonRoundTripTest(unittest.TestCase):
    def test_decimal_precision_survives_json(self) -> None:
        s = MonthlyWeekSummary(
            employees=[_emp("ALICE", 1, "X", basic="40.33", mf="0.50", ss="0.25", ah="8.00")],
            employee_totals=[_tot("X", basic="40.33", mf="0.50", ss="0.25", ah="8.00")],
            adjustments=[MonthlyAdjustment("ALICE", "Basic Hours", Decimal("-1.50"))],
            start_date="01/04/2026",
            end_date="07/04/2026",
        )
        s.non_agency_total = Decimal("49.08")
        s.grouped_totals = {"X": _tot("X", basic="40.33", mf="0.50", ss="0.25", ah="8.00")}
        blob = monthly_summaries_to_json([s])
        loaded = monthly_summaries_from_json(blob)[0]
        self.assertEqual(loaded.employees[0].BasicHours, Decimal("40.33"))
        self.assertEqual(loaded.employees[0].TotalPaidHours, Decimal("49.08"))
        self.assertEqual(loaded.employee_totals[0].BasicHours, Decimal("40.33"))
        self.assertEqual(loaded.adjustments[0].Value, Decimal("-1.50"))
        self.assertEqual(loaded.non_agency_total, Decimal("49.08"))
        self.assertEqual(loaded.grouped_totals["X"].BasicHours, Decimal("40.33"))

    def test_json_blob_is_json_safe(self) -> None:
        import json
        s = MonthlyWeekSummary(employees=[_emp("ALICE", 1, "X", basic="1.50")])
        blob = monthly_summaries_to_json([s])
        # Must round-trip through stdlib json (Django sessions use json by default).
        text = json.dumps(blob)
        self.assertIn('"1.50"', text)
        self.assertEqual(json.loads(text), blob)


class AgencyNonHourlyTest(unittest.TestCase):
    def test_agency_employee_marked_non_hourly_raises(self) -> None:
        s = MonthlyWeekSummary(
            employees=[_emp("ALICE", 501, "A-EL CLNR", basic="40")],
            employee_totals=[_tot("A-EL CLNR", basic="40")],
        )
        with self.assertRaisesRegex(ValueError, "Agency employee cannot be non-hourly"):
            build_monthly_excel_bytes([s], non_hourly_names={"ALICE"})

    def test_non_agency_non_hourly_is_allowed(self) -> None:
        s = MonthlyWeekSummary(
            employees=[_emp("ALICE", 501, "D-PROD", basic="40")],
            employee_totals=[_tot("D-PROD", basic="40")],
        )
        out = build_monthly_excel_bytes([s], non_hourly_names={"ALICE"})
        self.assertTrue(out)


class BuildExcelStructureTest(unittest.TestCase):
    def _two_week_input(self) -> list[MonthlyWeekSummary]:
        # Two weeks, one employee in D-PROD plus one in agency, plus one non-hourly admin.
        s1 = MonthlyWeekSummary(
            employees=[
                _emp("ALICE", 501, "D-PROD", basic="40.00", mf="2.00", ss="0", ah="0"),
                _emp("BOB", 502, "A-EL CLNR", basic="20.00", mf="0", ss="0", ah="0"),
                _emp("ADMIN1", 999, "Z-OFFICE", basic="35.00", mf="0", ss="0", ah="0", hourly=False),
            ],
            employee_totals=[
                _tot("D-PROD", basic="40.00", mf="2.00"),
                _tot("A-EL CLNR", basic="20.00"),
                _tot("Z-OFFICE", basic="35.00"),
            ],
            start_date="01/04/2026",
            end_date="07/04/2026",
        )
        s1.non_agency_total = Decimal("75.00") + Decimal("2.00")
        s2 = MonthlyWeekSummary(
            employees=[
                _emp("ALICE", 501, "D-PROD", basic="38.50", mf="1.50", ss="0", ah="0"),
                _emp("BOB", 502, "A-EL CLNR", basic="22.00", mf="0", ss="0", ah="0"),
                _emp("ADMIN1", 999, "Z-OFFICE", basic="35.00", mf="0", ss="0", ah="0", hourly=False),
            ],
            employee_totals=[
                _tot("D-PROD", basic="38.50", mf="1.50"),
                _tot("A-EL CLNR", basic="22.00"),
                _tot("Z-OFFICE", basic="35.00"),
            ],
            start_date="08/04/2026",
            end_date="14/04/2026",
        )
        s2.non_agency_total = Decimal("73.50") + Decimal("1.50")
        # Pre-compute per-week grouped totals (parse_monthly_week_file would normally do this).
        for s in (s1, s2):
            grouped: dict[str, MonthlyEmployeeTotal] = {}
            for t in s.employee_totals:
                k = _grouped_key(t.Category)
                if k not in grouped:
                    grouped[k] = MonthlyEmployeeTotal(Category=k)
                g = grouped[k]
                g.BasicHours += t.BasicHours
                g.MonFriOvertime += t.MonFriOvertime
                g.SatSunOvertime += t.SatSunOvertime
                g.AnnualHoliday += t.AnnualHoliday
                g.TotalPaidHours += t.TotalPaidHours
            s.grouped_totals = grouped
        return [s1, s2]

    def test_workbook_has_week_sheets_and_summary(self) -> None:
        data = build_monthly_excel_bytes(self._two_week_input(), non_hourly_names={"ADMIN1"})
        with zipfile.ZipFile(BytesIO(data)) as zf:
            wbxml = zf.read("xl/workbook.xml").decode("utf-8")
        for name in ("Week1", "Week2", "Summary"):
            self.assertIn(name, wbxml, msg=f"missing sheet {name!r}")

    def test_summary_employee_totals_match_sum_of_weeks(self) -> None:
        data = build_monthly_excel_bytes(self._two_week_input(), non_hourly_names={"ADMIN1"})
        wb = load_workbook(BytesIO(data))
        ws = wb["Summary"]
        # Find ALICE's row in the Summary employees block (starting row 4).
        found = False
        for r in range(4, 20):
            name = ws.cell(r, 1).value
            if name == "ALICE":
                # Decimal('40.00') + Decimal('38.50') = 78.50, OT = 2.0 + 1.5 = 3.5
                self.assertEqual(Decimal(str(ws.cell(r, 4).value)), Decimal("78.50"))
                self.assertEqual(Decimal(str(ws.cell(r, 5).value)), Decimal("3.50"))
                found = True
                break
        self.assertTrue(found, "ALICE row not found on Summary")

    def test_summary_has_non_hourly_subtraction_row_for_admin_category(self) -> None:
        data = build_monthly_excel_bytes(self._two_week_input(), non_hourly_names={"ADMIN1"})
        wb = load_workbook(BytesIO(data))
        ws = wb["Summary"]
        # Walk all cells looking for the marker text.
        marker = None
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value == "Z-OFFICE non-hourly hours":
                    marker = cell
                    break
            if marker:
                break
        self.assertIsNotNone(marker, "expected 'Z-OFFICE non-hourly hours' subtraction row on Summary")
        # Subtraction row writes negative basic hours equal to admin's combined basic (35+35=70).
        basic = Decimal(str(ws.cell(marker.row, 4).value))
        self.assertEqual(basic, Decimal("-70.00"))

    def test_non_agency_footer_excludes_a_dash_categories(self) -> None:
        data = build_monthly_excel_bytes(self._two_week_input(), non_hourly_names={"ADMIN1"})
        wb = load_workbook(BytesIO(data))
        ws = wb["Summary"]
        # Walk the workbook and find the last row with a numeric col D (BasicHours) not in a labelled row.
        last_numeric_row = None
        for r in range(ws.max_row, 1, -1):
            v = ws.cell(r, 4).value
            label = ws.cell(r, 2).value
            if v is not None and label in (None, ""):
                last_numeric_row = r
                break
        self.assertIsNotNone(last_numeric_row)
        # Non-agency footer = sum of D-PROD (78.5) + Z-OFFICE (70) basic = 148.5; A-EL CLNR (42) excluded.
        self.assertEqual(Decimal(str(ws.cell(last_numeric_row, 4).value)), Decimal("148.50"))


class WeekFileRoundTripTest(unittest.TestCase):
    """End-to-end: build a fake Weekly-format xlsx, then parse_monthly_week_file reads it back."""

    def _make_week_xlsx(self) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "Week 1")
        ws.cell(1, 4, "Start Date")
        ws.cell(1, 5, "D 01/04/2026")
        ws.cell(1, 7, "End Date")
        ws.cell(1, 8, "D 07/04/2026")
        # Headers row 3
        ws.cell(3, 1, "Name")
        ws.cell(3, 2, "Category")
        ws.cell(3, 3, "SageNo")
        ws.cell(3, 4, "BasicHours")
        ws.cell(3, 5, "MonFriOvertime")
        ws.cell(3, 6, "SatSunOvertime")
        ws.cell(3, 7, "AnnualHoliday")
        ws.cell(3, 8, "TotalPaidHours")
        # Two employees row 4-5
        ws.cell(4, 1, "ALICE")
        ws.cell(4, 2, "D-PROD")
        ws.cell(4, 3, 501)
        ws.cell(4, 4, 40.5)
        ws.cell(4, 5, 2.25)
        ws.cell(4, 6, 0)
        ws.cell(4, 7, 0)
        ws.cell(4, 8, 42.75)
        ws.cell(5, 1, "BOB")
        ws.cell(5, 2, "A-EL CLNR")
        ws.cell(5, 3, 502)
        ws.cell(5, 4, 38.0)
        ws.cell(5, 5, 0)
        ws.cell(5, 6, 0)
        ws.cell(5, 7, 0)
        ws.cell(5, 8, 38.0)
        # 3 blank rows then Adjustments header at row 9 (matches Weekly.exe writer offset).
        ws.cell(9, 2, "Adjustments - 1")
        ws.cell(10, 2, "Name")
        ws.cell(10, 3, "Type")
        ws.cell(10, 4, "Value")
        ws.cell(11, 2, "ALICE")
        ws.cell(11, 3, "Basic Hours")
        ws.cell(11, 4, "-1.50")
        # Totals: blank row, then Category header at row 13.
        ws.cell(13, 2, "Category")
        ws.cell(13, 4, "BasicHours")
        ws.cell(13, 5, "MonFriOvertime")
        ws.cell(13, 6, "SatSunOvertime")
        ws.cell(13, 7, "AnnualHoliday")
        ws.cell(13, 8, "TotalPaidHours")
        ws.cell(14, 2, "D-PROD")
        ws.cell(14, 4, 40.5)
        ws.cell(14, 5, 2.25)
        ws.cell(14, 6, 0)
        ws.cell(14, 7, 0)
        ws.cell(14, 8, 42.75)
        ws.cell(15, 2, "A-EL CLNR")
        ws.cell(15, 4, 38.0)
        ws.cell(15, 5, 0)
        ws.cell(15, 6, 0)
        ws.cell(15, 7, 0)
        ws.cell(15, 8, 38.0)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_parse_strips_d_prefix_from_dates(self) -> None:
        s = parse_monthly_week_file(self._make_week_xlsx())
        self.assertEqual(s.start_date, "01/04/2026")
        self.assertEqual(s.end_date, "07/04/2026")

    def test_parse_employees_as_decimal(self) -> None:
        s = parse_monthly_week_file(self._make_week_xlsx())
        self.assertEqual(len(s.employees), 2)
        a = s.employees[0]
        self.assertEqual(a.Name, "ALICE")
        self.assertEqual(a.SageNo, 501)
        self.assertEqual(a.BasicHours, Decimal("40.5"))
        self.assertEqual(a.TotalPaidHours, Decimal("42.75"))
        self.assertIsInstance(a.BasicHours, Decimal)

    def test_parse_adjustments_skips_subheader_row(self) -> None:
        s = parse_monthly_week_file(self._make_week_xlsx())
        self.assertEqual(len(s.adjustments), 1)
        self.assertEqual(s.adjustments[0].Name, "ALICE")
        self.assertEqual(s.adjustments[0].Type, "Basic Hours")
        self.assertEqual(s.adjustments[0].Value, Decimal("-1.50"))

    def test_parse_non_agency_total_excludes_a_prefix(self) -> None:
        s = parse_monthly_week_file(self._make_week_xlsx())
        # D-PROD 42.75 is included; A-EL CLNR 38.0 is excluded.
        self.assertEqual(s.non_agency_total, Decimal("42.75"))

    def test_parse_grouped_totals_split_for_agency(self) -> None:
        s = parse_monthly_week_file(self._make_week_xlsx())
        # "A-EL CLNR" should bucket under "A-EL CLNR"; "D-PROD" under "D-PR".
        self.assertIn("A-EL CLNR", s.grouped_totals)
        self.assertIn("D-PR", s.grouped_totals)
        self.assertEqual(s.grouped_totals["D-PR"].TotalPaidHours, Decimal("42.75"))


if __name__ == "__main__":
    unittest.main()
